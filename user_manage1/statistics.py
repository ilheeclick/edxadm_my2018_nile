# -*- coding: utf-8 -*-
from django.db import connection
from openpyxl import load_workbook
import statistics_query
from django.template import Context
from django.http import HttpResponse
from django.template.loader import get_template
from pymongo import MongoClient
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import os
from operator import itemgetter
import datetime
from openpyxl.styles import Alignment

from management.settings import EXCEL_PATH, dic_univ, database_id, debug
from time import gmtime, strftime

# 일일통계
def statistics_excel(request, date):

    if debug: print 'Step 1',strftime("%Y-%m-%d %H:%M:%S", gmtime())
    # Get course name
    course_ids_all = statistics_query.course_ids_all()

    client = MongoClient(database_id, 27017)
    db = client.edxapp
    pb = ''
    ov = ''
    org = ''
    courseOrgs = {}
    courseNames = {}

    if debug: print 'Step 2',strftime("%Y-%m-%d %H:%M:%S", gmtime())
    for c in course_ids_all:
        cid = str(c[0])
        courseId = cid

        if debug: print 'check >>> ', courseId, cid.split('+')[0], cid.split('+')[1], cid.split('+')[2]

        cid = courseId.split('+')[1]
        run = courseId.split('+')[2]

        cursor = db.modulestore.active_versions.find({'course':cid, 'run': run})
        for document in cursor:
            pb = document.get('versions').get('published-branch')
            org = document.get('org')
            break
        cursor.close()

        cursor = db.modulestore.structures.find({'_id':pb})
        for document in cursor:
            blocks = document.get('blocks')
            for block in blocks:
                blocktype = block.get('block_type')
                if blocktype == 'course':
                    fields = block.get('fields')
                    for field in fields:
                        if field == 'display_name':
                            if debug: print 'field :', field
                            dn = fields['display_name']
                            if debug: print 'dn1 :', dn
                            if dn == '' or dn is None:
                                if debug: print 'dn is empty'

                                ov = document.get('original_version')

                                if debug: print 'ov :', ov
                                cursor2 = db.modulestore.structures.find({'_id':ov})
                                for document2 in cursor2:
                                    blocks2 = document2.get('blocks')
                                    for block2 in blocks2:
                                        fields2 = block2.get('fields')
                                        dn = fields2['display_name']
                                        if debug: print 'dn2 :', dn
                                        break
                                    break
                                cursor2.close()

                            courseOrgs[courseId] = org
                            courseNames[courseId] = dn

                            print 'display_name = ', courseId, pb,  dn.encode('utf8')

                            break
                        else:
                            continue
                    break
                else:
                    continue
            break
        cursor.close()

    if debug: print 'Step 3',strftime("%Y-%m-%d %H:%M:%S", gmtime())
    thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

    if debug: print 'query1',strftime("%Y-%m-%d %H:%M:%S", gmtime())
    user_join_new = statistics_query.user_join_new(date)
    if debug: print 'query2',strftime("%Y-%m-%d %H:%M:%S", gmtime())
    user_join_total = statistics_query.user_join_total(date)
    if debug: print 'query3',strftime("%Y-%m-%d %H:%M:%S", gmtime())
    course_count_distinct = statistics_query.course_count_distinct(date)
    if debug: print 'query4',strftime("%Y-%m-%d %H:%M:%S", gmtime())
    course_count_new = statistics_query.course_count_new(date)
    if debug: print 'query5',strftime("%Y-%m-%d %H:%M:%S", gmtime())
    course_count_total = statistics_query.course_count_total(date)
    if debug: print 'query6',strftime("%Y-%m-%d %H:%M:%S", gmtime())
    edu_new = statistics_query.edu_new(date)
    if debug: print 'query7',strftime("%Y-%m-%d %H:%M:%S", gmtime())
    edu_total = statistics_query.edu_total(date)
    if debug: print 'query8',strftime("%Y-%m-%d %H:%M:%S", gmtime())
    age_new = statistics_query.age_new(date)
    if debug: print 'query9',strftime("%Y-%m-%d %H:%M:%S", gmtime())
    age_total = statistics_query.age_total(date)
    if debug: print 'query10',strftime("%Y-%m-%d %H:%M:%S", gmtime())
    age_edu = statistics_query.age_edu(date)
    if debug: print 'query11',strftime("%Y-%m-%d %H:%M:%S", gmtime())
    course_user = statistics_query.course_user(date)
    if debug: print 'query12',strftime("%Y-%m-%d %H:%M:%S", gmtime())
    course_univ = statistics_query.course_univ(date)
    if debug: print 'query13',strftime("%Y-%m-%d %H:%M:%S", gmtime())
    course_user_total =statistics_query.course_user_total(date)
    if debug: print 'query14',strftime("%Y-%m-%d %H:%M:%S", gmtime())
    course_univ_total =statistics_query.course_univ_total(date)
    if debug: print 'query15',strftime("%Y-%m-%d %H:%M:%S", gmtime())
    course_age = statistics_query.course_age(date)
    if debug: print 'query16',strftime("%Y-%m-%d %H:%M:%S", gmtime())
    course_edu = statistics_query.course_edu(date)

    if debug: print 'Step 4',strftime("%Y-%m-%d %H:%M:%S", gmtime())

    saveName = 'K-Mooc'+date+'.xlsx'
    savePath = EXCEL_PATH + saveName




    if os.path.isfile(savePath) and not debug:
        if debug: print 'pass'
        pass

    else:
        if debug: print 'make'
        wb = load_workbook(EXCEL_PATH + 'basic.xlsx')
        ws1 = wb['user_count']
        ws2 = wb['course_count']
        ws3 = wb['course_count_total']
        ws4 = wb['course_age']
        ws5 = wb['course_edu']

        #가입현황
        ws1['B4'] = user_join_new
        ws1['C4'] = user_join_total
        ws1['D4'] = course_count_new
        ws1['E4'] = course_count_total
        ws1['F4'] = course_count_distinct

        #연령구분
        sort = [
            (9,0),
            (10,1),
            (11,2),
            (12,3),
            (13,4),
            (14,5)
        ]
        if debug: print 'ws1 start',strftime("%Y-%m-%d %H:%M:%S", gmtime())
        for (number, number1) in sort:
            ws1['C' + str(number)] = age_new[number1][0]
            ws1['D' + str(number)] = age_new[number1][1]

        for (number, number1) in sort:
            ws1['F' + str(number)] = age_total[number1][0]
            ws1['G' + str(number)] = age_total[number1][1]

        #학력구분
        sort = [
            (20,0),
            (21,1),
            (22,2),
            (23,3),
            (24,4),
            (25,5),
            (26,6),
            (27,7),
            (28,8)
        ]

        for (number, number1) in sort:
            ws1['C' + str(number)] = edu_new[number1][0]
            ws1['D' + str(number)] = edu_new[number1][1]

        for (number, number1) in sort:
            ws1['F' + str(number)] = edu_total[number1][0]
            ws1['G' + str(number)] = edu_total[number1][1]

        #연령학력
        sort = [
            (34,0),
            (35,1),
            (36,2),
            (37,3),
            (38,4),
            (39,5)
        ]

        for (number, number1) in sort:
            ws1['C' + str(number)] = age_edu[number1][0]
            ws1['D' + str(number)] = age_edu[number1][1]
            ws1['E' + str(number)] = age_edu[number1][2]
            ws1['F' + str(number)] = age_edu[number1][3]
            ws1['G' + str(number)] = age_edu[number1][4]
            ws1['H' + str(number)] = age_edu[number1][5]
            ws1['I' + str(number)] = age_edu[number1][6]
            ws1['J' + str(number)] = age_edu[number1][7]
            ws1['K' + str(number)] = age_edu[number1][8]
        if debug: print 'ws1 end',strftime("%Y-%m-%d %H:%M:%S", gmtime())


        #======================================================================================================================================================


        if debug: print 'ws2 start',strftime("%Y-%m-%d %H:%M:%S", gmtime())
        #코스별 수강자 # LJH수정
        # sorted(courseInfo.items(), key=itemgetter(1))
        rn1 = 3
        rn2 = 0



        sortlist = list()
        for c in course_user:
            if str(courseOrgs[c[0]]) in dic_univ:
                c = (dic_univ[str(courseOrgs[c[0]])],courseNames[c[0]], ) + c
            else:
                c = (str(courseOrgs[c[0]]),courseNames[c[0]], ) + c
            print '0================================'
            print c
            print '================================='
            sortlist.append(c)

        sortlist.sort(key=itemgetter(0,4,1))
        for s in sortlist:
            orgName = s[0]
            cName = s[1]
            cId1 = s[3]
            cId2 = s[4]
            cnt = s[5]

            ws2['A' + str(rn1)] = orgName
            ws2['B' + str(rn1)] = cName
            ws2['C' + str(rn1)] = cId1
            ws2['D' + str(rn1)] = cId2
            ws2['E' + str(rn1)] = cnt
            # set border
            ws2['A' + str(rn1)].border = thin_border
            ws2['B' + str(rn1)].border = thin_border
            ws2['C' + str(rn1)].border= thin_border
            ws2['D' + str(rn1)].border = thin_border
            ws2['E' + str(rn1)].border = thin_border
            rn1 += 1
            rn2 += 1

        rn1 = 3

        # for univ, cnt in course_univ:
        sortlist = list()
        for c in course_univ:
            if c[0] in dic_univ:
                c = (dic_univ[c[0]],) + c
            else:
                c = (c[0], ) + c

            print '1================================'
            print c
            print '================================='
            sortlist.append(c)

        sortlist.sort(key=itemgetter(0))

        startCharNo1 = 72 # H
        startCharNo2 = 65 # A
        positionChar = ''
        isExpension = False

        for s in sortlist:
            orgName = s[0]
            cnt = s[2]

            if startCharNo1 > 90:
                startCharNo1 = 65

                if isExpension:
                    startCharNo2 += 1
                else:
                    isExpension = True

            if not isExpension:
                positionChar = chr(startCharNo1)
            else:
                positionChar = chr(startCharNo2) + chr(startCharNo1)

            print 'positionChar:', positionChar



            ws2[positionChar + '2'] = orgName
            ws2[positionChar + '3'] = cnt

            ws2[positionChar + '2'].alignment = Alignment(horizontal="center")

            # set border
            ws2[positionChar + '2'].border = thin_border
            ws2[positionChar + '3'].border = thin_border



            # print cId, cName
            # H 열부터 횡으로 증가. Z 까지 갔을경우 AA 로 다시 시작

            # 기존에 열으로 추가되는 로직
            # ----------------------------------------------

            # ws2['G' + str(rn1)] = orgName
            # ws2['H' + str(rn1)] = cnt
            # # set border
            # ws2['G' + str(rn1)].border = thin_border
            # ws2['H' + str(rn1)].border = thin_border
            # rn1 += 1

            # ----------------------------------------------

            startCharNo1 += 1

        if debug: print 'ws2 end',strftime("%Y-%m-%d %H:%M:%S", gmtime())

        if debug: print 'ws3 start',strftime("%Y-%m-%d %H:%M:%S", gmtime())


        #코스별 수강자 누적
        # sorted(courseInfo.items(), key=itemgetter(1))
        rn1 = 3
        rn2 = 0
        # for cId, cId1, cId2, cnt in course_user_total:
        sortlist = list()
        for c in course_user_total:
            if str(courseOrgs[c[0]]) in dic_univ:
                c = (dic_univ[str(courseOrgs[c[0]])],courseNames[c[0]], ) + c
            else:
                c = (str(courseOrgs[c[0]]),courseNames[c[0]], ) + c
            print '2================================'
            print c
            print '================================='
            sortlist.append(c)

        sortlist.sort(key=itemgetter(0,4,1))
        for s in sortlist:
            orgName = s[0]
            cName = s[1]
            cId1 = s[3]
            cId2 = s[4]
            cnt = s[5]
            # print cId, cName

            ws3['A' + str(rn1)] = orgName
            ws3['B' + str(rn1)] = cName
            ws3['C' + str(rn1)] = cId1
            ws3['D' + str(rn1)] = cId2
            ws3['E' + str(rn1)] = cnt
            # set border
            ws3['A' + str(rn1)].border = thin_border
            ws3['B' + str(rn1)].border = thin_border
            ws3['C' + str(rn1)].border = thin_border
            ws3['D' + str(rn1)].border = thin_border
            ws3['E' + str(rn1)].border = thin_border
            rn1 += 1
            rn2 += 1

        rn1 = 3
        # for univ, cnt in course_univ_total:
        sortlist = list()
        for c in course_univ_total:
            if c[0] in dic_univ:
                c = (dic_univ[c[0]],) + c
            else:
                c = (c[0], ) + c
            print '3================================'
            print c
            print '================================='
            sortlist.append(c)

        sortlist.sort(key=itemgetter(0))

        startCharNo1 = 72 # H
        startCharNo2 = 65 # A
        positionChar = ''
        isExpension = False

        for s in sortlist:
            orgName = s[0]
            cnt = s[2]

            if startCharNo1 > 90:
                startCharNo1 = 65

                if isExpension:
                    startCharNo2 += 1
                else:
                    isExpension = True

            if not isExpension:
                positionChar = chr(startCharNo1)
            else:
                positionChar = chr(startCharNo2) + chr(startCharNo1)

            print 'positionChar:', positionChar

            ws3[positionChar + '2'] = orgName
            ws3[positionChar + '3'] = cnt

            ws2[positionChar + '2'].alignment = Alignment(horizontal="center")

            # set border
            ws3[positionChar + '2'].border = thin_border
            ws3[positionChar + '3'].border = thin_border

            startCharNo1 += 1
        if debug: print 'ws3 end',strftime("%Y-%m-%d %H:%M:%S", gmtime())


        if debug: print 'ws4 start',strftime("%Y-%m-%d %H:%M:%S", gmtime())

        #코스별 연령
        rn1 = 3
        # for org, cId, cId1, cId2, age1, age2, age3, age4, age5, age6 in course_age:
        sortlist = list()
        for c in course_age:
            if c[0] in dic_univ:
                c = (dic_univ[c[0]],) + c
            else:
                c = (c[0],) + c
            print '4================================'
            print c
            print '================================='

            sortlist.append(c)

        sortlist.sort(key=itemgetter(0,4,1))
        for s in sortlist:
            orgName = s[0]
            cName = courseNames[s[2]]
            cId1 = s[3]
            cId2 = s[4]
            age1 = s[5]
            age2 = s[6]
            age3 = s[7]
            age4 = s[8]
            age5 = s[9]
            age6 = s[10]

            ws4['A' + str(rn1)] = orgName
            ws4['B' + str(rn1)] = cName
            ws4['C' + str(rn1)] = cId1
            ws4['D' + str(rn1)] = cId2
            ws4['E' + str(rn1)] = age1
            ws4['F' + str(rn1)] = age2
            ws4['G' + str(rn1)] = age3
            ws4['H' + str(rn1)] = age4
            ws4['I' + str(rn1)] = age5
            ws4['J' + str(rn1)] = age6

            # set border
            ws4['A' + str(rn1)].border = thin_border
            ws4['B' + str(rn1)].border = thin_border
            ws4['C' + str(rn1)].border = thin_border
            ws4['D' + str(rn1)].border = thin_border
            ws4['E' + str(rn1)].border = thin_border
            ws4['F' + str(rn1)].border = thin_border
            ws4['G' + str(rn1)].border = thin_border
            ws4['H' + str(rn1)].border = thin_border
            ws4['I' + str(rn1)].border = thin_border
            ws4['J' + str(rn1)].border = thin_border
            rn1 += 1
        if debug: print 'ws4 end',strftime("%Y-%m-%d %H:%M:%S", gmtime())

        if debug: print 'ws5 start',strftime("%Y-%m-%d %H:%M:%S", gmtime())

        #코스별 학력
        rn1 = 3
        # for org, cId, cId1, cId2, edu1, edu2, edu3, edu4, edu5, edu6, edu7, edu8, edu9 in course_edu:
        sortlist = list()
        for c in course_edu:
            if c[0] in dic_univ:
                c = (dic_univ[c[0]],) + c
            else:
                c = (c[0],) + c
            print '5================================'
            print c
            print '================================='
            sortlist.append(c)

        sortlist.sort(key=itemgetter(0,4,1))
        for s in sortlist:
            orgName = s[0]
            cName = courseNames[s[2]]
            cId1 = s[3]
            cId2 = s[4]
            edu1 = s[5]
            edu2 = s[6]
            edu3 = s[7]
            edu4 = s[8]
            edu5 = s[9]
            edu6 = s[10]
            edu7 = s[11]
            edu8 = s[12]
            edu9 = s[13]
            # print cId, cName

            ws5['A' + str(rn1)] = orgName

            ws5['B' + str(rn1)] = cName
            ws5['C' + str(rn1)] = cId1
            ws5['D' + str(rn1)] = cId2
            ws5['E' + str(rn1)] = edu1
            ws5['F' + str(rn1)] = edu2
            ws5['G' + str(rn1)] = edu3
            ws5['H' + str(rn1)] = edu4
            ws5['I' + str(rn1)] = edu5
            ws5['J' + str(rn1)] = edu6
            ws5['K' + str(rn1)] = edu7
            ws5['L' + str(rn1)] = edu8
            ws5['M' + str(rn1)] = edu9

            # set border
            ws5['A' + str(rn1)].border = thin_border
            ws5['B' + str(rn1)].border = thin_border
            ws5['C' + str(rn1)].border = thin_border
            ws5['D' + str(rn1)].border = thin_border
            ws5['E' + str(rn1)].border = thin_border
            ws5['F' + str(rn1)].border = thin_border
            ws5['G' + str(rn1)].border = thin_border
            ws5['H' + str(rn1)].border = thin_border
            ws5['I' + str(rn1)].border = thin_border
            ws5['J' + str(rn1)].border = thin_border
            ws5['K' + str(rn1)].border = thin_border
            ws5['L' + str(rn1)].border = thin_border
            ws5['M' + str(rn1)].border = thin_border
            rn1 += 1


        if debug: print 'ws2 end',strftime("%Y-%m-%d %H:%M:%S", gmtime())

        wb.save(savePath)
        if debug: print 'Step 5',strftime("%Y-%m-%d %H:%M:%S", gmtime())
    return HttpResponse('/manage/static/excel/' + saveName, content_type='application/vnd.ms-excel')



def certificate_excel(request, courseId):

    print 'courseId', courseId

    d = datetime.date.today()
    year = d.year
    month = d.month
    day = d.day

    if month < 10:
        month = '0' + str(month)
    if day < 10:
        day = '0' + str(day)


    print 'month', month
    print 'day', day

    certificates = statistics_query.certificateInfo(courseId)

    courseName = ''
    pb = ''
    ov = ''

    client = MongoClient(database_id, 27017)
    db = client.edxapp

    wb = load_workbook(EXCEL_PATH + '/basic_cert.xlsx')

    for c in certificates:
        cid = str(c[2])

        print 'cid', cid

        cursor = db.modulestore.active_versions.find({'course':cid})
        for document in cursor:
            print '>> 1'
            pb = document.get('versions').get('published-branch')
            break
        cursor.close()

        cursor = db.modulestore.structures.find({'_id':pb})
        for document in cursor:
            print '>> 2'
            ov = document.get('original_version')
            break
        cursor.close()

        cursor = db.modulestore.structures.find({'_id':ov})
        for document in cursor:
            print '>> 3'
            blocks = document.get('blocks')
            for block in blocks:
                print '>> 4'
                fields = block.get('fields')
                for field in fields:
                    print '>> 5'
                    dn = fields['display_name']
                    courseName = dn
                    break
                break
            break
        break
        cursor.close()

    thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

    # dic_status = {'KHUk':u'경희대학교', 'KoreaUnivK':u'고려대학교', 'PNUk':u'부산대학교', 'SNUk':u'서울대학교', 'SKKUk':u'성균관대학교',
    #             'YSUk':u'연세대학교', 'EwhaK':u'이화여자대학교', 'POSTECHk':u'포항공과대학교', 'KAISTk':u'한국과학기술원', 'HYUk':u'한양대학교'}


    ws1 = wb['certificates']

    row = 2
    for c in certificates:
        print '0------------------------------------------------------------------'
        print c
        print '-------------------------------------------------------------------'

        ws1['A' + str(row)] = dic_univ[c[0]]
        ws1['B' + str(row)] = courseName
        ws1['C' + str(row)] = c[2]
        ws1['D' + str(row)] = c[3]
        ws1['E' + str(row)] = c[4]
        ws1['F' + str(row)] = c[5]

        ws1['A' + str(row)].border = thin_border
        ws1['B' + str(row)].border = thin_border
        ws1['C' + str(row)].border = thin_border
        ws1['D' + str(row)].border = thin_border
        ws1['E' + str(row)].border = thin_border
        ws1['F' + str(row)].border = thin_border
        row += 1

    saveName = 'K-Mooc_certificate_' + str(cid)  + '_' + str(year) + str(month) + str(day) +'.xlsx'
    savePath = '/Users/redukyo/workspace/management3/static/excel/' + saveName

    wb.save(savePath)

    return HttpResponse('/manage/static/excel/' + saveName, content_type='application/vnd.ms-excel')

def statistics_excel3(request, date):

    saveName = 'K-MoocMonth'+date+'.xlsx'
    savePath = EXCEL_PATH + saveName

    if os.path.isfile(savePath) and not debug:
        print '@@@ statistics_excel3 make pass'
        pass
    else:
        print '@@@ statistics_excel3 make working'
        member_statistics = statistics_query.member_statistics(date)
        country_statistics = statistics_query.country_statistics(date)

        wb = load_workbook(EXCEL_PATH + 'basic_month.xlsx')
        thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

        COUNTRIES = {
            "AF": "Afghanistan",
            "AX": "Åland Islands",
            "AL": "Albania",
            "DZ": "Algeria",
            "AS": "American Samoa",
            "AD": "Andorra",
            "AO": "Angola",
            "AI": "Anguilla",
            "AQ": "Antarctica",
            "AG": "Antigua and Barbuda",
            "AR": "Argentina",
            "AM": "Armenia",
            "AW": "Aruba",
            "AU": "Australia",
            "AT": "Austria",
            "AZ": "Azerbaijan",
            "BS": "Bahamas",
            "BH": "Bahrain",
            "BD": "Bangladesh",
            "BB": "Barbados",
            "BY": "Belarus",
            "BE": "Belgium",
            "BZ": "Belize",
            "BJ": "Benin",
            "BM": "Bermuda",
            "BT": "Bhutan",
            "BO": "Bolivia (Plurinational State of)",
            "BQ": "Bonaire, Sint Eustatius and Saba",
            "BA": "Bosnia and Herzegovina",
            "BW": "Botswana",
            "BV": "Bouvet Island",
            "BR": "Brazil",
            "IO": "British Indian Ocean Territory",
            "BN": "Brunei Darussalam",
            "BG": "Bulgaria",
            "BF": "Burkina Faso",
            "BI": "Burundi",
            "CV": "Cabo Verde",
            "KH": "Cambodia",
            "CM": "Cameroon",
            "CA": "Canada",
            "KY": "Cayman Islands",
            "CF": "Central African Republic",
            "TD": "Chad",
            "CL": "Chile",
            "CN": "China",
            "CX": "Christmas Island",
            "CC": "Cocos (Keeling) Islands",
            "CO": "Colombia",
            "KM": "Comoros",
            "CD": "Congo (the Democratic Republic of the)",
            "CG": "Congo",
            "CK": "Cook Islands",
            "CR": "Costa Rica",
            "CI": "Côte d'Ivoire",
            "HR": "Croatia",
            "CU": "Cuba",
            "CW": "Curaçao",
            "CY": "Cyprus",
            "CZ": "Czech Republic",
            "DK": "Denmark",
            "DJ": "Djibouti",
            "DM": "Dominica",
            "DO": "Dominican Republic",
            "EC": "Ecuador",
            "EG": "Egypt",
            "SV": "El Salvador",
            "GQ": "Equatorial Guinea",
            "ER": "Eritrea",
            "EE": "Estonia",
            "ET": "Ethiopia",
            "FK": "Falkland Islands  [Malvinas]",
            "FO": "Faroe Islands",
            "FJ": "Fiji",
            "FI": "Finland",
            "FR": "France",
            "GF": "French Guiana",
            "PF": "French Polynesia",
            "TF": "French Southern Territories",
            "GA": "Gabon",
            "GM": "Gambia",
            "GE": "Georgia",
            "DE": "Germany",
            "GH": "Ghana",
            "GI": "Gibraltar",
            "GR": "Greece",
            "GL": "Greenland",
            "GD": "Grenada",
            "GP": "Guadeloupe",
            "GU": "Guam",
            "GT": "Guatemala",
            "GG": "Guernsey",
            "GN": "Guinea",
            "GW": "Guinea-Bissau",
            "GY": "Guyana",
            "HT": "Haiti",
            "HM": "Heard Island and McDonald Islands",
            "VA": "Holy See",
            "HN": "Honduras",
            "HK": "Hong Kong",
            "HU": "Hungary",
            "IS": "Iceland",
            "IN": "India",
            "ID": "Indonesia",
            "IR": "Iran (Islamic Republic of)",
            "IQ": "Iraq",
            "IE": "Ireland",
            "IM": "Isle of Man",
            "IL": "Israel",
            "IT": "Italy",
            "JM": "Jamaica",
            "JP": "Japan",
            "JE": "Jersey",
            "JO": "Jordan",
            "KZ": "Kazakhstan",
            "KE": "Kenya",
            "KI": "Kiribati",
            "KP": "Korea (the Democratic People's Republic of)",
            "KR": "Korea (the Republic of)",
            "KW": "Kuwait",
            "KG": "Kyrgyzstan",
            "LA": "Lao People's Democratic Republic",
            "LV": "Latvia",
            "LB": "Lebanon",
            "LS": "Lesotho",
            "LR": "Liberia",
            "LY": "Libya",
            "LI": "Liechtenstein",
            "LT": "Lithuania",
            "LU": "Luxembourg",
            "MO": "Macao",
            "MK": "Macedonia (the former Yugoslav Republic of)",
            "MG": "Madagascar",
            "MW": "Malawi",
            "MY": "Malaysia",
            "MV": "Maldives",
            "ML": "Mali",
            "MT": "Malta",
            "MH": "Marshall Islands",
            "MQ": "Martinique",
            "MR": "Mauritania",
            "MU": "Mauritius",
            "YT": "Mayotte",
            "MX": "Mexico",
            "FM": "Micronesia (Federated States of)",
            "MD": "Moldova (the Republic of)",
            "MC": "Monaco",
            "MN": "Mongolia",
            "ME": "Montenegro",
            "MS": "Montserrat",
            "MA": "Morocco",
            "MZ": "Mozambique",
            "MM": "Myanmar",
            "NA": "Namibia",
            "NR": "Nauru",
            "NP": "Nepal",
            "NL": "Netherlands",
            "NC": "New Caledonia",
            "NZ": "New Zealand",
            "NI": "Nicaragua",
            "NE": "Niger",
            "NG": "Nigeria",
            "NU": "Niue",
            "NF": "Norfolk Island",
            "MP": "Northern Mariana Islands",
            "NO": "Norway",
            "OM": "Oman",
            "PK": "Pakistan",
            "PW": "Palau",
            "PS": "Palestine, State of",
            "PA": "Panama",
            "PG": "Papua New Guinea",
            "PY": "Paraguay",
            "PE": "Peru",
            "PH": "Philippines",
            "PN": "Pitcairn",
            "PL": "Poland",
            "PT": "Portugal",
            "PR": "Puerto Rico",
            "QA": "Qatar",
            "RE": "Réunion",
            "RO": "Romania",
            "RU": "Russian Federation",
            "RW": "Rwanda",
            "BL": "Saint Barthélemy",
            "SH": "Saint Helena, Ascension and Tristan da Cunha",
            "KN": "Saint Kitts and Nevis",
            "LC": "Saint Lucia",
            "MF": "Saint Martin (French part)",
            "PM": "Saint Pierre and Miquelon",
            "VC": "Saint Vincent and the Grenadines",
            "WS": "Samoa",
            "SM": "San Marino",
            "ST": "Sao Tome and Principe",
            "SA": "Saudi Arabia",
            "SN": "Senegal",
            "RS": "Serbia",
            "SC": "Seychelles",
            "SL": "Sierra Leone",
            "SG": "Singapore",
            "SX": "Sint Maarten (Dutch part)",
            "SK": "Slovakia",
            "SI": "Slovenia",
            "SB": "Solomon Islands",
            "SO": "Somalia",
            "ZA": "South Africa",
            "GS": "South Georgia and the South Sandwich Islands",
            "SS": "South Sudan",
            "ES": "Spain",
            "LK": "Sri Lanka",
            "SD": "Sudan",
            "SR": "Suriname",
            "SJ": "Svalbard and Jan Mayen",
            "SZ": "Swaziland",
            "SE": "Sweden",
            "CH": "Switzerland",
            "SY": "Syrian Arab Republic",
            "TW": "Taiwan (Province of China)",
            "TJ": "Tajikistan",
            "TZ": "Tanzania, United Republic of",
            "TH": "Thailand",
            "TL": "Timor-Leste",
            "TG": "Togo",
            "TK": "Tokelau",
            "TO": "Tonga",
            "TT": "Trinidad and Tobago",
            "TN": "Tunisia",
            "TR": "Turkey",
            "TM": "Turkmenistan",
            "TC": "Turks and Caicos Islands",
            "TV": "Tuvalu",
            "UG": "Uganda",
            "UA": "Ukraine",
            "AE": "United Arab Emirates",
            "GB": "United Kingdom of Great Britain and Northern Ireland",
            "UM": "United States Minor Outlying Islands",
            "US": "United States of America",
            "UY": "Uruguay",
            "UZ": "Uzbekistan",
            "VU": "Vanuatu",
            "VE": "Venezuela (Bolivarian Republic of)",
            "VN": "Viet Nam",
            "VG": "Virgin Islands (British)",
            "VI": "Virgin Islands (U.S.)",
            "WF": "Wallis and Futuna",
            "EH": "Western Sahara",
            "YE": "Yemen",
            "ZM": "Zambia",
            "ZW": "Zimbabwe",
        }

        ws1 = wb['Sheet1']

        row = 2
        for c in member_statistics:
            ws1['B' + str(row+1)] = c[0]
            ws1['B' + str(row+2)] = c[1]
            ws1['B' + str(row+3)] = c[2]

        row = 8
        for c in country_statistics:

            ws1['A' + str(row)] = c[0]
            ws1['B' + str(row)] = c[1]
            ws1['C' + str(row)] = COUNTRIES[c[0]]

            ws1['A' + str(row)].border = thin_border
            ws1['B' + str(row)].border = thin_border
            ws1['C' + str(row)].border = thin_border
            row += 1

        wb.save(savePath)

    return HttpResponse('/manage/static/excel/' + saveName, content_type='application/vnd.ms-excel')



