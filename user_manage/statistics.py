# -*- coding: utf-8 -*-
from django.db import connection
from openpyxl import load_workbook
import statistics_query
from django.template import Context
from django.http import HttpResponse
from django.template.loader import get_template
from pymongo import MongoClient
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Style
import os
from operator import itemgetter
import datetime

def statistics_excel(request, date):




    # Get course name
    course_ids_all = statistics_query.course_ids_all()

    client = MongoClient('192.168.1.113', 27017)
    db = client.edxapp
    pb = ''
    ov = ''
    org = ''
    courseOrgs = {}
    courseNames = {}

    for c in course_ids_all:
        cid = str(c[0])
        courseId = cid
        cid = cid.split('+')[1]

        # print '1. cid = ', cid

        cursor = db.modulestore.active_versions.find({'course':cid})
        for document in cursor:
            pb = document.get('versions').get('published-branch')
            org = document.get('org')
            break
        cursor.close()

        cursor = db.modulestore.structures.find({'_id':pb})
        for document in cursor:
            ov = document.get('original_version')
            # print '2. original_vertion = ', ov
            break
        cursor.close()

        cursor = db.modulestore.structures.find({'_id':ov})
        for document in cursor:
            blocks = document.get('blocks')
            # print 'size = ', len(blocks)
            for block in blocks:
                    fields = block.get('fields')
                    for field in fields:
                            dn = fields['display_name']
                            # print '3. display_name = ', dn.encode('utf8')
                            courseOrgs[courseId] = org
                            courseNames[courseId] = dn
                            break
                    break
            break
        cursor.close()

    thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

    dic_univ = {'KHUk':u'경희대학교', 'KoreaUnivK':u'고려대학교', 'PNUk':u'부산대학교', 'SNUk':u'서울대학교', 'SKKUk':u'성균관대학교',
                'YSUk':u'연세대학교', 'EwhaK':u'이화여자대학교', 'POSTECHk':u'포항공과대학교', 'KAISTk':u'한국과학기술원', 'HYUk':u'한양대학교'}

    # print '======'
    #for univ in sortlist:
    #    print str(univ)
    # print '======'


    user_join_new = statistics_query.user_join_new(date)
    user_join_total = statistics_query.user_join_total(date)
    course_count_distinct = statistics_query.course_count_distinct(date)
    course_count_new = statistics_query.course_count_new(date)
    course_count_total = statistics_query.course_count_total(date)
    edu_new = statistics_query.edu_new(date)
    edu_total = statistics_query.edu_total(date)
    age_new = statistics_query.age_new(date)
    age_total = statistics_query.age_total(date)
    age_edu = statistics_query.age_edu(date)
    course_user = statistics_query.course_user(date)
    course_univ = statistics_query.course_univ(date)
    course_user_total =statistics_query.course_user_total(date)
    course_univ_total =statistics_query.course_univ_total(date)
    course_age = statistics_query.course_age(date)
    course_edu = statistics_query.course_edu(date)

    saveName = 'K-Mooc'+date+'.xlsx'
    savePath = '/home/project/management/static/excel/' + saveName

    if os.path.isfile(savePath) and False:
        pass

    else:
        wb = load_workbook('/home/project/management/static/excel/basic.xlsx')
        ws1 = wb['user_count']
        ws2 = wb['course_count']
        ws3 = wb['course_count_total']
        ws4 = wb['course_age']
        ws5 = wb['course_edu']

        #가입현황
        ws1['B4'] = user_join_new
        ws1['C4'] = user_join_total
        ws1['D4'] = course_count_distinct
        ws1['E4'] = course_count_new
        ws1['F4'] = course_count_total

        #학력구분
        sort = [(9,0),(10,1),(11,2),(12,3),(13,4),
                (14,5),(15,6),(16,7),(17,8)]

        for (number, number1) in sort:
            ws1['C' + str(number)] = edu_new[number1][0]
            ws1['D' + str(number)] = edu_new[number1][1]

        for (number, number1) in sort:
            ws1['F' + str(number)] = edu_total[number1][0]
            ws1['G' + str(number)] = edu_total[number1][1]

        #연령구분
        sort = [(23,0),(24,1),(25,2),(26,3),(27,4),(28,5)]

        for (number, number1) in sort:
            ws1['C' + str(number)] = age_new[number1][0]
            ws1['D' + str(number)] = age_new[number1][1]

        for (number, number1) in sort:
            ws1['F' + str(number)] = age_total[number1][0]
            ws1['G' + str(number)] = age_total[number1][1]

        #연령학력
        sort = [(34,0),(35,1),(36,2),(37,3),(38,4),(39,5)]

        for (number, number1) in sort:
            ws1['C' + str(number)] = age_edu[number1][0]
            ws1['D' + str(number)] = age_edu[number1][1]
            ws1['E' + str(number)] = age_edu[number1][2]
            ws1['F' + str(number)] = age_edu[number1][3]
            ws1['G' + str(number)] = age_edu[number1][4]
            ws1['H' + str(number)] = age_edu[number1][5]
            ws1['I' + str(number)] = age_edu[number1][6]
            ws1['J' + str(number)] = age_edu[number1][7]
            ws1['K' + str(number)] = age_edu[number1][8]


        #======================================================================================================================================================


        #코스별 수강자 # LJH수정
        # sorted(courseInfo.items(), key=itemgetter(1))
        rn1 = 3
        rn2 = 0


        sortlist = list()
        for c in course_user:
            if str(courseOrgs[c[0]]) in dic_univ:
                c = (dic_univ[str(courseOrgs[c[0]])],courseNames[c[0]], ) + c
            else:
                c = (str(courseOrgs[c[0]]),courseNames[c[0]], ) + c
            print '0================================'
            print c
            print '================================='
            sortlist.append(c)

        sortlist.sort(key=itemgetter(0,4,1))
        for s in sortlist:
            orgName = s[0]
            cName = s[1]
            cId1 = s[3]
            cId2 = s[4]
            cnt = s[5]

            ws2['A' + str(rn1)] = orgName
            ws2['B' + str(rn1)] = cName
            ws2['C' + str(rn1)] = cId1
            ws2['D' + str(rn1)] = cId2
            ws2['E' + str(rn1)] = cnt
            # set border
            ws2['A' + str(rn1)].border = thin_border
            ws2['B' + str(rn1)].border = thin_border
            ws2['C' + str(rn1)].border = thin_border
            ws2['D' + str(rn1)].border = thin_border
            ws2['E' + str(rn1)].border = thin_border
            rn1 += 1
            rn2 += 1

        rn1 = 3

        # for univ, cnt in course_univ:
        sortlist = list()
        for c in course_univ:
            if c[0] in dic_univ:
                c = (dic_univ[c[0]],) + c
            else:
                c = (c[0], ) + c
            print '1================================'
            print c
            print '================================='
            sortlist.append(c)

        sortlist.sort(key=itemgetter(0))
        for s in sortlist:
            orgName = s[0]
            cnt = s[2]

            # print cId, cName

            ws2['G' + str(rn1)] = orgName
            ws2['H' + str(rn1)] = cnt
            # set border
            ws2['G' + str(rn1)].border = thin_border
            ws2['H' + str(rn1)].border = thin_border
            rn1 += 1


        #코스별 수강자 누적
        # sorted(courseInfo.items(), key=itemgetter(1))
        rn1 = 3
        rn2 = 0
        # for cId, cId1, cId2, cnt in course_user_total:
        sortlist = list()
        for c in course_user_total:
            if str(courseOrgs[c[0]]) in dic_univ:
                c = (dic_univ[str(courseOrgs[c[0]])],courseNames[c[0]], ) + c
            else:
                c = (str(courseOrgs[c[0]]),courseNames[c[0]], ) + c
            print '2================================'
            print c
            print '================================='
            sortlist.append(c)

        sortlist.sort(key=itemgetter(0,4,1))
        for s in sortlist:
            orgName = s[0]
            cName = s[1]
            cId1 = s[3]
            cId2 = s[4]
            cnt = s[5]
            # print cId, cName

            ws3['A' + str(rn1)] = orgName
            ws3['B' + str(rn1)] = cName
            ws3['C' + str(rn1)] = cId1
            ws3['D' + str(rn1)] = cId2
            ws3['E' + str(rn1)] = cnt
            # set border
            ws3['A' + str(rn1)].border = thin_border
            ws3['B' + str(rn1)].border = thin_border
            ws3['C' + str(rn1)].border = thin_border
            ws3['D' + str(rn1)].border = thin_border
            ws3['E' + str(rn1)].border = thin_border
            rn1 += 1
            rn2 += 1

        rn1 = 3
        # for univ, cnt in course_univ_total:
        sortlist = list()
        for c in course_univ_total:
            if c[0] in dic_univ:
                c = (dic_univ[c[0]],) + c
            else:
                c = (c[0], ) + c
            print '3================================'
            print c
            print '================================='
            sortlist.append(c)

        sortlist.sort(key=itemgetter(0))
        for s in sortlist:
            orgName = s[0]
            cnt = s[2]
            # print cId, cName

            ws3['G' + str(rn1)] = orgName
            ws3['H' + str(rn1)] = cnt
            # set border
            ws3['G' + str(rn1)].border = thin_border
            ws3['H' + str(rn1)].border = thin_border
            rn1 += 1

        #코스별 연령
        rn1 = 3
        # for org, cId, cId1, cId2, age1, age2, age3, age4, age5, age6 in course_age:
        sortlist = list()
        for c in course_age:
            if c[0] in dic_univ:
                c = (dic_univ[c[0]],) + c
            else:
                c = (c[0],) + c
            print '4================================'
            print c
            print '================================='

            sortlist.append(c)

        sortlist.sort(key=itemgetter(0,4,1))
        for s in sortlist:
            orgName = s[0]
            cName = courseNames[s[2]]
            cId1 = s[3]
            cId2 = s[4]
            age1 = s[5]
            age2 = s[6]
            age3 = s[7]
            age4 = s[8]
            age5 = s[9]
            age6 = s[10]

            ws4['A' + str(rn1)] = orgName
            ws4['B' + str(rn1)] = cName
            ws4['C' + str(rn1)] = cId1
            ws4['D' + str(rn1)] = cId2
            ws4['E' + str(rn1)] = age1
            ws4['F' + str(rn1)] = age2
            ws4['G' + str(rn1)] = age3
            ws4['H' + str(rn1)] = age4
            ws4['I' + str(rn1)] = age5
            ws4['J' + str(rn1)] = age6

            # set border
            ws4['A' + str(rn1)].border = thin_border
            ws4['B' + str(rn1)].border = thin_border
            ws4['C' + str(rn1)].border = thin_border
            ws4['D' + str(rn1)].border = thin_border
            ws4['E' + str(rn1)].border = thin_border
            ws4['F' + str(rn1)].border = thin_border
            ws4['G' + str(rn1)].border = thin_border
            ws4['H' + str(rn1)].border = thin_border
            ws4['I' + str(rn1)].border = thin_border
            ws4['J' + str(rn1)].border = thin_border
            rn1 += 1


        #코스별 학력
        rn1 = 3
        # for org, cId, cId1, cId2, edu1, edu2, edu3, edu4, edu5, edu6, edu7, edu8, edu9 in course_edu:
        sortlist = list()
        for c in course_edu:
            if c[0] in dic_univ:
                c = (dic_univ[c[0]],) + c
            else:
                c = (c[0],) + c
            print '5================================'
            print c
            print '================================='
            sortlist.append(c)

        sortlist.sort(key=itemgetter(0,4,1))
        for s in sortlist:
            orgName = s[0]
            cName = courseNames[s[2]]
            cId1 = s[3]
            cId2 = s[4]
            edu1 = s[5]
            edu2 = s[6]
            edu3 = s[7]
            edu4 = s[8]
            edu5 = s[9]
            edu6 = s[10]
            edu7 = s[11]
            edu8 = s[12]
            edu9 = s[13]
            # print cId, cName

            ws5['A' + str(rn1)] = orgName

            ws5['B' + str(rn1)] = cName
            ws5['C' + str(rn1)] = cId1
            ws5['D' + str(rn1)] = cId2
            ws5['E' + str(rn1)] = edu1
            ws5['F' + str(rn1)] = edu2
            ws5['G' + str(rn1)] = edu3
            ws5['H' + str(rn1)] = edu4
            ws5['I' + str(rn1)] = edu5
            ws5['J' + str(rn1)] = edu6
            ws5['K' + str(rn1)] = edu7
            ws5['L' + str(rn1)] = edu8
            ws5['M' + str(rn1)] = edu9

            # set border
            ws5['A' + str(rn1)].border = thin_border
            ws5['B' + str(rn1)].border = thin_border
            ws5['C' + str(rn1)].border = thin_border
            ws5['D' + str(rn1)].border = thin_border
            ws5['E' + str(rn1)].border = thin_border
            ws5['F' + str(rn1)].border = thin_border
            ws5['G' + str(rn1)].border = thin_border
            ws5['H' + str(rn1)].border = thin_border
            ws5['I' + str(rn1)].border = thin_border
            ws5['J' + str(rn1)].border = thin_border
            ws5['K' + str(rn1)].border = thin_border
            ws5['L' + str(rn1)].border = thin_border
            ws5['M' + str(rn1)].border = thin_border
            rn1 += 1


        wb.save(savePath)

    return HttpResponse('/manage/static/excel/' + saveName, content_type='application/vnd.ms-excel')


def certificate_excel(request, courseId):

    print 'courseId', courseId

    d = datetime.date.today()
    year = d.year
    month = d.month
    day = d.day

    if month < 10:
        month = '0' + str(month)
    if day < 10:
        day = '0' + str(day)


    print 'month', month
    print 'day', day

    certificates = statistics_query.certificateInfo(courseId)

    courseName = ''
    pb = ''
    ov = ''

    client = MongoClient('192.168.1.113', 27017)
    db = client.edxapp

    wb = load_workbook('/home/project/management/static/excel/basic_cert.xlsx')

    for c in certificates:
        cid = str(c[2])

        print 'cid', cid

        cursor = db.modulestore.active_versions.find({'course':cid})
        for document in cursor:
            print '>> 1'
            pb = document.get('versions').get('published-branch')
            break
        cursor.close()

        cursor = db.modulestore.structures.find({'_id':pb})
        for document in cursor:
            print '>> 2'
            ov = document.get('original_version')
            break
        cursor.close()

        cursor = db.modulestore.structures.find({'_id':ov})
        for document in cursor:
            print '>> 3'
            blocks = document.get('blocks')
            for block in blocks:
                print '>> 4'
                fields = block.get('fields')
                for field in fields:
                    print '>> 5'
                    dn = fields['display_name']
                    courseName = dn
                    break
                break
            break
        break
        cursor.close()

    thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

    dic_univ = {'KHUk':u'경희대학교', 'KoreaUnivK':u'고려대학교', 'PNUk':u'부산대학교', 'SNUk':u'서울대학교', 'SKKUk':u'성균관대학교',
                'YSUk':u'연세대학교', 'EwhaK':u'이화여자대학교', 'POSTECHk':u'포항공과대학교', 'KAISTk':u'한국과학기술원', 'HYUk':u'한양대학교'}

    # dic_status = {'KHUk':u'경희대학교', 'KoreaUnivK':u'고려대학교', 'PNUk':u'부산대학교', 'SNUk':u'서울대학교', 'SKKUk':u'성균관대학교',
    #             'YSUk':u'연세대학교', 'EwhaK':u'이화여자대학교', 'POSTECHk':u'포항공과대학교', 'KAISTk':u'한국과학기술원', 'HYUk':u'한양대학교'}


    ws1 = wb['certificates']

    row = 2
    for c in certificates:
        print '0------------------------------------------------------------------'
        print c
        print '-------------------------------------------------------------------'

        ws1['A' + str(row)] = dic_univ[c[0]]
        ws1['B' + str(row)] = courseName
        ws1['C' + str(row)] = c[2]
        ws1['D' + str(row)] = c[3]
        ws1['E' + str(row)] = c[4]
        ws1['F' + str(row)] = c[5]

        ws1['A' + str(row)].border = thin_border
        ws1['B' + str(row)].border = thin_border
        ws1['C' + str(row)].border = thin_border
        ws1['D' + str(row)].border = thin_border
        ws1['E' + str(row)].border = thin_border
        ws1['F' + str(row)].border = thin_border
        row += 1

    saveName = 'K-Mooc_certificate_' + str(cid)  + '_' + str(year) + str(month) + str(day) +'.xlsx'
    savePath = '/home/project/management/static/excel/' + saveName

    wb.save(savePath)

    return HttpResponse('/manage/static/excel/' + saveName, content_type='application/vnd.ms-excel')

def statistics_excel3(request, date):

    member_statistics = statistics_query.member_statistics(date)
    country_statistics = statistics_query.country_statistics(date)

    wb = load_workbook('/home/project/management/static/excel/basic_month.xlsx')
    thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
    
    COUNTRIES = {
        "AF": "Afghanistan",
        "AX": "Åland Islands",
        "AL": "Albania",
        "DZ": "Algeria",
        "AS": "American Samoa",
        "AD": "Andorra",
        "AO": "Angola",
        "AI": "Anguilla",
        "AQ": "Antarctica",
        "AG": "Antigua and Barbuda",
        "AR": "Argentina",
        "AM": "Armenia",
        "AW": "Aruba",
        "AU": "Australia",
        "AT": "Austria",
        "AZ": "Azerbaijan",
        "BS": "Bahamas",
        "BH": "Bahrain",
        "BD": "Bangladesh",
        "BB": "Barbados",
        "BY": "Belarus",
        "BE": "Belgium",
        "BZ": "Belize",
        "BJ": "Benin",
        "BM": "Bermuda",
        "BT": "Bhutan",
        "BO": "Bolivia (Plurinational State of)",
        "BQ": "Bonaire, Sint Eustatius and Saba",
        "BA": "Bosnia and Herzegovina",
        "BW": "Botswana",
        "BV": "Bouvet Island",
        "BR": "Brazil",
        "IO": "British Indian Ocean Territory",
        "BN": "Brunei Darussalam",
        "BG": "Bulgaria",
        "BF": "Burkina Faso",
        "BI": "Burundi",
        "CV": "Cabo Verde",
        "KH": "Cambodia",
        "CM": "Cameroon",
        "CA": "Canada",
        "KY": "Cayman Islands",
        "CF": "Central African Republic",
        "TD": "Chad",
        "CL": "Chile",
        "CN": "China",
        "CX": "Christmas Island",
        "CC": "Cocos (Keeling) Islands",
        "CO": "Colombia",
        "KM": "Comoros",
        "CD": "Congo (the Democratic Republic of the)",
        "CG": "Congo",
        "CK": "Cook Islands",
        "CR": "Costa Rica",
        "CI": "Côte d'Ivoire",
        "HR": "Croatia",
        "CU": "Cuba",
        "CW": "Curaçao",
        "CY": "Cyprus",
        "CZ": "Czech Republic",
        "DK": "Denmark",
        "DJ": "Djibouti",
        "DM": "Dominica",
        "DO": "Dominican Republic",
        "EC": "Ecuador",
        "EG": "Egypt",
        "SV": "El Salvador",
        "GQ": "Equatorial Guinea",
        "ER": "Eritrea",
        "EE": "Estonia",
        "ET": "Ethiopia",
        "FK": "Falkland Islands  [Malvinas]",
        "FO": "Faroe Islands",
        "FJ": "Fiji",
        "FI": "Finland",
        "FR": "France",
        "GF": "French Guiana",
        "PF": "French Polynesia",
        "TF": "French Southern Territories",
        "GA": "Gabon",
        "GM": "Gambia",
        "GE": "Georgia",
        "DE": "Germany",
        "GH": "Ghana",
        "GI": "Gibraltar",
        "GR": "Greece",
        "GL": "Greenland",
        "GD": "Grenada",
        "GP": "Guadeloupe",
        "GU": "Guam",
        "GT": "Guatemala",
        "GG": "Guernsey",
        "GN": "Guinea",
        "GW": "Guinea-Bissau",
        "GY": "Guyana",
        "HT": "Haiti",
        "HM": "Heard Island and McDonald Islands",
        "VA": "Holy See",
        "HN": "Honduras",
        "HK": "Hong Kong",
        "HU": "Hungary",
        "IS": "Iceland",
        "IN": "India",
        "ID": "Indonesia",
        "IR": "Iran (Islamic Republic of)",
        "IQ": "Iraq",
        "IE": "Ireland",
        "IM": "Isle of Man",
        "IL": "Israel",
        "IT": "Italy",
        "JM": "Jamaica",
        "JP": "Japan",
        "JE": "Jersey",
        "JO": "Jordan",
        "KZ": "Kazakhstan",
        "KE": "Kenya",
        "KI": "Kiribati",
        "KP": "Korea (the Democratic People's Republic of)",
        "KR": "Korea (the Republic of)",
        "KW": "Kuwait",
        "KG": "Kyrgyzstan",
        "LA": "Lao People's Democratic Republic",
        "LV": "Latvia",
        "LB": "Lebanon",
        "LS": "Lesotho",
        "LR": "Liberia",
        "LY": "Libya",
        "LI": "Liechtenstein",
        "LT": "Lithuania",
        "LU": "Luxembourg",
        "MO": "Macao",
        "MK": "Macedonia (the former Yugoslav Republic of)",
        "MG": "Madagascar",
        "MW": "Malawi",
        "MY": "Malaysia",
        "MV": "Maldives",
        "ML": "Mali",
        "MT": "Malta",
        "MH": "Marshall Islands",
        "MQ": "Martinique",
        "MR": "Mauritania",
        "MU": "Mauritius",
        "YT": "Mayotte",
        "MX": "Mexico",
        "FM": "Micronesia (Federated States of)",
        "MD": "Moldova (the Republic of)",
        "MC": "Monaco",
        "MN": "Mongolia",
        "ME": "Montenegro",
        "MS": "Montserrat",
        "MA": "Morocco",
        "MZ": "Mozambique",
        "MM": "Myanmar",
        "NA": "Namibia",
        "NR": "Nauru",
        "NP": "Nepal",
        "NL": "Netherlands",
        "NC": "New Caledonia",
        "NZ": "New Zealand",
        "NI": "Nicaragua",
        "NE": "Niger",
        "NG": "Nigeria",
        "NU": "Niue",
        "NF": "Norfolk Island",
        "MP": "Northern Mariana Islands",
        "NO": "Norway",
        "OM": "Oman",
        "PK": "Pakistan",
        "PW": "Palau",
        "PS": "Palestine, State of",
        "PA": "Panama",
        "PG": "Papua New Guinea",
        "PY": "Paraguay",
        "PE": "Peru",
        "PH": "Philippines",
        "PN": "Pitcairn",
        "PL": "Poland",
        "PT": "Portugal",
        "PR": "Puerto Rico",
        "QA": "Qatar",
        "RE": "Réunion",
        "RO": "Romania",
        "RU": "Russian Federation",
        "RW": "Rwanda",
        "BL": "Saint Barthélemy",
        "SH": "Saint Helena, Ascension and Tristan da Cunha",
        "KN": "Saint Kitts and Nevis",
        "LC": "Saint Lucia",
        "MF": "Saint Martin (French part)",
        "PM": "Saint Pierre and Miquelon",
        "VC": "Saint Vincent and the Grenadines",
        "WS": "Samoa",
        "SM": "San Marino",
        "ST": "Sao Tome and Principe",
        "SA": "Saudi Arabia",
        "SN": "Senegal",
        "RS": "Serbia",
        "SC": "Seychelles",
        "SL": "Sierra Leone",
        "SG": "Singapore",
        "SX": "Sint Maarten (Dutch part)",
        "SK": "Slovakia",
        "SI": "Slovenia",
        "SB": "Solomon Islands",
        "SO": "Somalia",
        "ZA": "South Africa",
        "GS": "South Georgia and the South Sandwich Islands",
        "SS": "South Sudan",
        "ES": "Spain",
        "LK": "Sri Lanka",
        "SD": "Sudan",
        "SR": "Suriname",
        "SJ": "Svalbard and Jan Mayen",
        "SZ": "Swaziland",
        "SE": "Sweden",
        "CH": "Switzerland",
        "SY": "Syrian Arab Republic",
        "TW": "Taiwan (Province of China)",
        "TJ": "Tajikistan",
        "TZ": "Tanzania, United Republic of",
        "TH": "Thailand",
        "TL": "Timor-Leste",
        "TG": "Togo",
        "TK": "Tokelau",
        "TO": "Tonga",
        "TT": "Trinidad and Tobago",
        "TN": "Tunisia",
        "TR": "Turkey",
        "TM": "Turkmenistan",
        "TC": "Turks and Caicos Islands",
        "TV": "Tuvalu",
        "UG": "Uganda",
        "UA": "Ukraine",
        "AE": "United Arab Emirates",
        "GB": "United Kingdom of Great Britain and Northern Ireland",
        "UM": "United States Minor Outlying Islands",
        "US": "United States of America",
        "UY": "Uruguay",
        "UZ": "Uzbekistan",
        "VU": "Vanuatu",
        "VE": "Venezuela (Bolivarian Republic of)",
        "VN": "Viet Nam",
        "VG": "Virgin Islands (British)",
        "VI": "Virgin Islands (U.S.)",
        "WF": "Wallis and Futuna",
        "EH": "Western Sahara",
        "YE": "Yemen",
        "ZM": "Zambia",
        "ZW": "Zimbabwe",
    }

    ws1 = wb['Sheet1']

    row = 2
    for c in member_statistics:
        ws1['B' + str(row+1)] = c[0]
        ws1['B' + str(row+2)] = c[1]
        ws1['B' + str(row+3)] = c[2]

    row = 8
    for c in country_statistics:

        ws1['A' + str(row)] = c[0]
        ws1['B' + str(row)] = c[1]
        ws1['C' + str(row)] = COUNTRIES[c[0]]

        ws1['A' + str(row)].border = thin_border
        ws1['B' + str(row)].border = thin_border
        ws1['C' + str(row)].border = thin_border
        row += 1

    saveName = 'K-MoocMonth'+date+'.xlsx'
    savePath = '/home/project/management/static/excel/' + saveName

    wb.save(savePath)

    return HttpResponse('/manage/static/excel/' + saveName, content_type='application/vnd.ms-excel')



