# -*- coding: utf-8 -*-

from openpyxl import load_workbook
import statistics_query
from django.http import HttpResponse
from pymongo import MongoClient
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import os
from operator import itemgetter
import datetime
from management.settings import EXCEL_PATH, dic_univ, database_id, debug, classfy, middle_classfy, countries
from bson.objectid import ObjectId
import time
import logging
import logging.handlers

'''
# 로거 인스턴스를 만든다
#logger = logging.get#logger('statistics log')

# 포매터를 만든다
fomatter = logging.Formatter('[%(levelname)s|%(filename)s:%(lineno)s] %(asctime)s > %(message)s')

# 스트림과 파일로 로그를 출력하는 핸들러를 각각 만든다.
fileHandler = logging.FileHandler('./statistics.log')
streamHandler = logging.StreamHandler()

# 각 핸들러에 포매터를 지정한다.
fileHandler.setFormatter(fomatter)
streamHandler.setFormatter(fomatter)

# 로거 인스턴스에 스트림 핸들러와 파일핸들러를 붙인다.
#logger.addHandler(fileHandler)
#logger.addHandler(streamHandler)

if debug:
    #logger.setLevel(logging.DEBUG)
'''


def style_base(cell):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center", vertical="center")


def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    try:
        top = Border(top=border.top)
        left = Border(left=border.left)
        right = Border(right=border.right)
        bottom = Border(bottom=border.bottom)

        first_cell = ws[cell_range.split(":")[0]]
        if alignment:
            ws.merge_cells(cell_range)
            first_cell.alignment = alignment

        rows = ws[cell_range]
        if font:
            first_cell.font = font

        for cell in rows[0]:
            cell.border = cell.border + top
        for cell in rows[-1]:
            cell.border = cell.border + bottom

        for row in rows:
            l = row[0]
            r = row[-1]
            l.border = l.border + left
            r.border = r.border + right
            if fill:
                for c in row:
                    c.fill = fill
    except Exception as e:
        print e


def certificate_excel(request, course_id):
    # print 'course_id', course_id

    d = datetime.date.today()
    year = d.year
    month = d.month
    day = d.day

    if month < 10:
        month = '0' + str(month)
    if day < 10:
        day = '0' + str(day)

    # print 'month', month
    # print 'day', day

    certificates = statistics_query.certificateInfo(course_id)

    courseName = ''
    pb = ''
    ov = ''

    client = MongoClient(database_id, 27017)
    db = client.edxapp

    wb = load_workbook(EXCEL_PATH + '/basic_cert.xlsx')

    for c in certificates:
        cid = str(c[2])

        # print 'cid', cid

        cursor = db.modulestore.active_versions.find({'course': cid})
        for document in cursor:
            # print '>> 1'
            pb = document.get('versions').get('published-branch')
            break
        cursor.close()

        cursor = db.modulestore.structures.find({'_id': pb})
        for document in cursor:
            # print '>> 2'
            ov = document.get('original_version')
            break
        cursor.close()

        cursor = db.modulestore.structures.find({'_id': ov})
        for document in cursor:
            # print '>> 3'
            blocks = document.get('blocks')
            for block in blocks:
                # print '>> 4'
                fields = block.get('fields')
                for field in fields:
                    # print '>> 5'
                    dn = fields['display_name']
                    courseName = dn
                    break
                break
            break
        break
        cursor.close()

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    ws1 = wb['certificates']

    row = 2
    for c in certificates:
        ws1['A' + str(row)] = dic_univ[c[0]]
        ws1['B' + str(row)] = courseName
        ws1['C' + str(row)] = c[2]
        ws1['D' + str(row)] = c[3]
        ws1['E' + str(row)] = c[4]
        ws1['F' + str(row)] = c[5]

        ws1['A' + str(row)].border = thin_border
        ws1['B' + str(row)].border = thin_border
        ws1['C' + str(row)].border = thin_border
        ws1['D' + str(row)].border = thin_border
        ws1['E' + str(row)].border = thin_border
        ws1['F' + str(row)].border = thin_border
        row += 1

    save_name = 'K-Mooc_certificate_' + str(cid) + '_' + str(year) + str(month) + str(day) + '.xlsx'
    save_path = EXCEL_PATH + save_name

    wb.save(save_path)

    return HttpResponse('/manage/static/excel/' + save_name, content_type='application/vnd.ms-excel')


# 일일통계
def statistics_excel(request, date):
    # print 'run statistics_excel'
    time.sleep(1)

    save_name = 'K-Mooc{0}.xlsx'.format(date)
    save_path = EXCEL_PATH + save_name

    if os.path.isfile(save_path):
        print 'file exists. so just download.'
    else:
        # Get course name
        course_ids_all = statistics_query.course_ids_all()

        print 'len(course_ids_all) = ', len(course_ids_all)

        client = MongoClient(database_id, 27017)
        db = client.edxapp
        course_orgs = {}
        course_classfys = {}
        course_middle_classfys = {}
        course_names = {}
        course_creates = {}
        course_enroll_starts = {}
        course_enroll_ends = {}
        course_starts = {}
        course_ends = {}
        course_edited = {}

        course_effort = {}
        course_video = {}
        course_week = {}
        course_cert_date = {}

        course_state = {}
        course_order = {}

        print 'step1 : course_info search'

        for course_id, display_name, course, org, start, end, enrollment_start, enrollment_end, effort, cert_date in course_ids_all:
            print course_id, org, display_name, type(start), start, type(end), end, type(enrollment_start), enrollment_start, type(enrollment_end), enrollment_end

            cid = course_id.split('+')[1]
            run = course_id.split('+')[2]

            # course_state setting
            utc_time = datetime.datetime.utcnow()

            if cert_date:
                course_state[course_id] = '이수증발급'
                course_order[course_id] = 1
            elif end and end < utc_time:
                course_state[course_id] = '종료'
                course_order[course_id] = 3
            elif start and start < utc_time < end:
                course_state[course_id] = '운영중'
                course_order[course_id] = 4
            elif start and utc_time < start:
                course_state[course_id] = '개강예정'
                course_order[course_id] = 5
            else:
                course_state[course_id] = '미정'
                course_order[course_id] = 9

            if cert_date:
                course_cert_date[course_id] = cert_date

            cursor = db.modulestore.active_versions.find_one({'course': cid, 'run': run})
            if not cursor:
                print 'not exists: ', cid, run
                continue

            pb = cursor.get('versions').get('published-branch')
            # course_orgs
            course_orgs[course_id] = cursor.get('org')
            course_creates[course_id] = cursor.get('edited_on')

            cursor = db.modulestore.structures.find_one({'_id': ObjectId(pb)}, {"blocks": {"$elemMatch": {"block_type": "course"}}})

            _classfy = cursor.get('blocks')[0].get('fields').get('classfy')  # classfy
            _mclassfy = cursor.get('blocks')[0].get('fields').get('middle_classfy')  # middle_classfy

            _course_edited = cursor.get('blocks')[0].get('edit_info').get('edited_on')  # middle_classfy

            if start is not None:
                course_starts[course_id] = start

            if end is not None:
                course_ends[course_id] = end

            if enrollment_start is not None:
                course_enroll_starts[course_id] = enrollment_start

            if enrollment_end is not None:
                course_enroll_ends[course_id] = enrollment_end

            if display_name is not None:
                course_names[course_id] = display_name

            if _classfy is not None and _classfy != 'all':
                course_classfys[course_id] = classfy[_classfy] if _classfy in classfy else _classfy

            if _mclassfy is not None and _mclassfy != 'all':
                course_middle_classfys[course_id] = middle_classfy[_mclassfy] if _mclassfy in middle_classfy else _mclassfy

            if _course_edited is not None:
                course_edited[course_id] = _course_edited

            if effort:
                course_effort[course_id] = effort.split('@')[0] if effort and '@' in effort else '-'
                course_week[course_id] = effort.split('@')[1].split('#')[0] if effort and '@' in effort and '#' in effort else '-'
                course_video[course_id] = effort.split('#')[1] if effort and '#' in effort else '-'

        print 'step2 : mysql search'

        # 요약
        auth_user_info = statistics_query.auth_user_info(date)
        student_courseenrollment_info = statistics_query.student_courseenrollment_info(date)
        certificate_info = statistics_query.overall_only_cert(date)

        # 회원가입/수강신청 세부사항
        overall_auth = statistics_query.overall_auth(date)
        overall_enroll = statistics_query.overall_enroll(date)
        overall_cert = statistics_query.overall_cert(date)

        # 연령구분
        # age_new = statistics_query.age_new(date)
        # age_total = statistics_query.age_total(date)

        # 학력구분
        # edu_new = statistics_query.edu_new(date)
        # edu_total = statistics_query.edu_total(date)

        # 연령별 학력
        # age_edu = statistics_query.age_edu(date)

        # 연령/성별
        age_gender_join = statistics_query.age_gender_join(date)
        age_gender_enroll = statistics_query.age_gender_enroll(date)
        age_gender_cert_half = statistics_query.age_gender_cert_half(date)
        age_gender_cert = statistics_query.age_gender_cert(date)

        # 학력/성별
        edu_gender_join = statistics_query.edu_gender_join(date)
        edu_gender_enroll = statistics_query.edu_gender_enroll(date)
        edu_gender_cert_half = statistics_query.edu_gender_cert_half(date)
        edu_gender_cert = statistics_query.edu_gender_cert(date)

        # 연령/학력
        age_edu_join = statistics_query.age_edu_join(date)
        age_edu_enroll = statistics_query.age_edu_enroll(date)
        age_edu_cert_half = statistics_query.age_edu_cert_half(date)
        age_edu_cert = statistics_query.age_edu_cert(date)

        print 'step3: info '
        wb = load_workbook(EXCEL_PATH + 'base.xlsx')

        ws1 = wb['overall']
        ws2 = wb['overall_demographic']
        ws3 = wb['by_course_KPI']
        ws4 = wb['by_course_demographic']

        # 20170816 추가 시트

        # excel style
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        # fill = PatternFill("solid", fgColor="eeeeee")
        fill = None

        font = Font(b=True, color="000000")
        # font = None
        al = Alignment(horizontal="center", vertical="center")

        # sheet1

        # 요약
        style_range(ws1, 'B2:C2', border=thin_border, fill=fill, font=font, alignment=al)
        style_range(ws1, 'D2:E2', border=thin_border, fill=fill, font=font, alignment=al)
        style_range(ws1, 'F2:G2', border=thin_border, fill=fill, font=font, alignment=al)

        # 세부사항
        style_range(ws1, 'B7:C8', border=thin_border, fill=fill, font=font, alignment=al)
        style_range(ws1, 'D7:E7', border=thin_border, fill=fill, font=font, alignment=al)
        style_range(ws1, 'F7:G7', border=thin_border, fill=fill, font=font, alignment=al)
        style_range(ws1, 'B9:B12', border=thin_border, fill=fill, font=font, alignment=al)
        style_range(ws1, 'B13:B15', border=thin_border, fill=fill, font=font, alignment=al)
        style_range(ws1, 'B16:B17', border=thin_border, fill=fill, font=font, alignment=al)

        # 연령/성별
        style_range(ws2, 'B2:C3', border=thin_border, fill=fill, font=font, alignment=al)
        style_range(ws2, 'D2:H2', border=thin_border, fill=fill, font=font, alignment=al)
        style_range(ws2, 'B4:B11', border=thin_border, fill=fill, font=font, alignment=al)
        style_range(ws2, 'B12:B19', border=thin_border, fill=fill, font=font, alignment=al)
        style_range(ws2, 'B20:B27', border=thin_border, fill=fill, font=font, alignment=al)
        style_range(ws2, 'B28:B35', border=thin_border, fill=fill, font=font, alignment=al)

        # 학력/성별
        style_range(ws2, 'B38:C39', border=thin_border, fill=fill, font=font, alignment=al)
        style_range(ws2, 'D38:H38', border=thin_border, fill=fill, font=font, alignment=al)
        style_range(ws2, 'B40:B49', border=thin_border, fill=fill, font=font, alignment=al)
        style_range(ws2, 'B50:B59', border=thin_border, fill=fill, font=font, alignment=al)
        style_range(ws2, 'B60:B69', border=thin_border, fill=fill, font=font, alignment=al)
        style_range(ws2, 'B70:B79', border=thin_border, fill=fill, font=font, alignment=al)

        # 연령/학력
        style_range(ws2, 'B82:C83', border=thin_border, fill=fill, font=font, alignment=al)
        style_range(ws2, 'D82:M82', border=thin_border, fill=fill, font=font, alignment=al)
        style_range(ws2, 'B84:B91', border=thin_border, fill=fill, font=font, alignment=al)
        style_range(ws2, 'B92:B99', border=thin_border, fill=fill, font=font, alignment=al)
        style_range(ws2, 'B100:B107', border=thin_border, fill=fill, font=font, alignment=al)
        style_range(ws2, 'B108:B115', border=thin_border, fill=fill, font=font, alignment=al)

        # by_course_KPI
        style_range(ws3, 'A1:K2', border=thin_border, fill=fill, font=font, alignment=al)
        style_range(ws3, 'L1:R2', border=thin_border, fill=fill, font=font, alignment=al)
        style_range(ws3, 'S1:W1', border=thin_border, fill=fill, font=font, alignment=al)
        style_range(ws3, 'S2:T2', border=thin_border, fill=fill, font=font, alignment=al)
        style_range(ws3, 'U2:V2', border=thin_border, fill=fill, font=font, alignment=al)
        style_range(ws3, 'X1:Y2', border=thin_border, fill=fill, font=font, alignment=al)
        style_range(ws3, 'Z1:Z2', border=thin_border, fill=fill, font=font, alignment=al)

        # by_course_demographic
        style_range(ws4, 'A1:K2', border=thin_border, fill=fill, font=font, alignment=al)
        style_range(ws4, 'L1:AE1', border=thin_border, fill=fill, font=font, alignment=al)
        style_range(ws4, 'AF1:AY1', border=thin_border, fill=fill, font=font, alignment=al)
        style_range(ws4, 'AZ1:BS1', border=thin_border, fill=fill, font=font, alignment=al)
        style_range(ws4, 'L2:O2', border=thin_border, fill=fill, font=font, alignment=al)
        style_range(ws4, 'P2:U2', border=thin_border, fill=fill, font=font, alignment=al)
        style_range(ws4, 'V2:AD2', border=thin_border, fill=fill, font=font, alignment=al)
        # style_range(ws4, 'AE2:AE3', border=thin_border, fill=fill, font=font, alignment=al)
        style_range(ws4, 'AF2:AI2', border=thin_border, fill=fill, font=font, alignment=al)
        style_range(ws4, 'AJ2:AO2', border=thin_border, fill=fill, font=font, alignment=al)
        style_range(ws4, 'AP2:AX2', border=thin_border, fill=fill, font=font, alignment=al)
        # style_range(ws4, 'AY2:AY3', border=thin_border, fill=fill, font=font, alignment=al)
        style_range(ws4, 'AZ2:BC2', border=thin_border, fill=fill, font=font, alignment=al)
        style_range(ws4, 'BD2:BI2', border=thin_border, fill=fill, font=font, alignment=al)
        style_range(ws4, 'BJ2:BR2', border=thin_border, fill=fill, font=font, alignment=al)
        # style_range(ws4, 'BS2:BS3', border=thin_border, fill=fill, font=font, alignment=al)

        # 가입현황
        # logger.info('가입현황')
        ws1['B4'] = auth_user_info[0][0]
        ws1['C4'] = auth_user_info[0][1]
        ws1['D4'] = student_courseenrollment_info[0][0]
        ws1['E4'] = student_courseenrollment_info[0][1]
        ws1['F4'] = certificate_info[0][0]
        ws1['G4'] = certificate_info[0][1]

        # 회원가입 / 수강신청 세부사항
        # logger.info('수강신청구분')
        ws1['E9'] = overall_auth[0][0]
        ws1['E10'] = overall_auth[0][1]
        ws1['E12'] = overall_auth[0][2]
        ws1['G9'] = overall_auth[0][3]
        ws1['G10'] = overall_auth[0][4]
        ws1['G12'] = overall_auth[0][5]

        #: 수강신청
        ws1['D13'] = overall_enroll[0][0]
        ws1['D14'] = overall_enroll[0][1]
        ws1['D15'] = overall_enroll[0][2]
        ws1['E13'] = overall_enroll[0][3]
        ws1['E14'] = overall_enroll[0][4]
        ws1['E15'] = overall_enroll[0][5]
        ws1['F13'] = overall_enroll[0][6]
        ws1['F14'] = overall_enroll[0][7]
        ws1['F15'] = overall_enroll[0][8]
        ws1['G13'] = overall_enroll[0][9]
        ws1['G14'] = overall_enroll[0][10]
        ws1['G15'] = overall_enroll[0][11]

        #: 이수
        ws1['D16'] = overall_cert[0][0]
        ws1['D17'] = overall_cert[0][1]
        ws1['E16'] = overall_cert[0][2]
        ws1['E17'] = overall_cert[0][3]
        ws1['F16'] = overall_cert[0][4]
        ws1['F17'] = overall_cert[0][5]
        ws1['G16'] = overall_cert[0][6]
        ws1['G17'] = overall_cert[0][7]

        # ::::::::::::::::::::::::::::::::::::::::::::::::::: 연령/성별

        start_row = 4
        for age_group, male, female, etc, no_gender in age_gender_join:
            ws2['D' + str(start_row)] = male
            ws2['E' + str(start_row)] = female
            ws2['F' + str(start_row)] = etc
            ws2['G' + str(start_row)] = no_gender
            start_row += 1

        start_row = 12
        for age_group, male, female, etc, no_gender in age_gender_enroll:
            ws2['D' + str(start_row)] = male
            ws2['E' + str(start_row)] = female
            ws2['F' + str(start_row)] = etc
            ws2['G' + str(start_row)] = no_gender
            start_row += 1

        start_row = 20
        for age_group, male, female, etc, no_gender in age_gender_cert_half:
            ws2['D' + str(start_row)] = male
            ws2['E' + str(start_row)] = female
            ws2['F' + str(start_row)] = etc
            ws2['G' + str(start_row)] = no_gender
            start_row += 1

        start_row = 28
        for age_group, male, female, etc, no_gender in age_gender_cert:
            ws2['D' + str(start_row)] = male
            ws2['E' + str(start_row)] = female
            ws2['F' + str(start_row)] = etc
            ws2['G' + str(start_row)] = no_gender
            start_row += 1

        # ::::::::::::::::::::::::::::::::::::::::::::::::::: 학력/성별

        start_row = 40
        for edu_group, male, female, etc, no_gender in edu_gender_join:
            ws2['D' + str(start_row)] = male
            ws2['E' + str(start_row)] = female
            ws2['F' + str(start_row)] = etc
            ws2['G' + str(start_row)] = no_gender
            start_row += 1

        start_row = 50
        for age_group, male, female, etc, no_gender in edu_gender_enroll:
            ws2['D' + str(start_row)] = male
            ws2['E' + str(start_row)] = female
            ws2['F' + str(start_row)] = etc
            ws2['G' + str(start_row)] = no_gender
            start_row += 1

        start_row = 60
        for age_group, male, female, etc, no_gender in edu_gender_cert_half:
            ws2['D' + str(start_row)] = male
            ws2['E' + str(start_row)] = female
            ws2['F' + str(start_row)] = etc
            ws2['G' + str(start_row)] = no_gender
            start_row += 1

        start_row = 70
        for age_group, male, female, etc, no_gender in edu_gender_cert:
            ws2['D' + str(start_row)] = male
            ws2['E' + str(start_row)] = female
            ws2['F' + str(start_row)] = etc
            ws2['G' + str(start_row)] = no_gender
            start_row += 1

        # ::::::::::::::::::::::::::::::::::::::::::::::::::: 연령/학력

        start_row = 84
        for age_group, edu1, edu2, edu3, edu4, edu5, edu6, edu7, edu8, edu9 in age_edu_join:
            ws2['D' + str(start_row)] = edu1
            ws2['E' + str(start_row)] = edu2
            ws2['F' + str(start_row)] = edu3
            ws2['G' + str(start_row)] = edu4
            ws2['H' + str(start_row)] = edu5
            ws2['I' + str(start_row)] = edu6
            ws2['J' + str(start_row)] = edu7
            ws2['K' + str(start_row)] = edu8
            ws2['L' + str(start_row)] = edu9
            start_row += 1

        start_row = 92
        for age_group, edu1, edu2, edu3, edu4, edu5, edu6, edu7, edu8, edu9 in age_edu_enroll:
            ws2['D' + str(start_row)] = edu1
            ws2['E' + str(start_row)] = edu2
            ws2['F' + str(start_row)] = edu3
            ws2['G' + str(start_row)] = edu4
            ws2['H' + str(start_row)] = edu5
            ws2['I' + str(start_row)] = edu6
            ws2['J' + str(start_row)] = edu7
            ws2['K' + str(start_row)] = edu8
            ws2['L' + str(start_row)] = edu9
            start_row += 1

        start_row = 100
        for age_group, edu1, edu2, edu3, edu4, edu5, edu6, edu7, edu8, edu9 in age_edu_cert_half:
            ws2['D' + str(start_row)] = edu1
            ws2['E' + str(start_row)] = edu2
            ws2['F' + str(start_row)] = edu3
            ws2['G' + str(start_row)] = edu4
            ws2['H' + str(start_row)] = edu5
            ws2['I' + str(start_row)] = edu6
            ws2['J' + str(start_row)] = edu7
            ws2['K' + str(start_row)] = edu8
            ws2['L' + str(start_row)] = edu9
            start_row += 1

        start_row = 108
        for age_group, edu1, edu2, edu3, edu4, edu5, edu6, edu7, edu8, edu9 in age_edu_cert:
            ws2['D' + str(start_row)] = edu1
            ws2['E' + str(start_row)] = edu2
            ws2['F' + str(start_row)] = edu3
            ws2['G' + str(start_row)] = edu4
            ws2['H' + str(start_row)] = edu5
            ws2['I' + str(start_row)] = edu6
            ws2['J' + str(start_row)] = edu7
            ws2['K' + str(start_row)] = edu8
            ws2['L' + str(start_row)] = edu9
            start_row += 1

        # :SHEET 3
        # ------------------------> by_course_KPI

        sortlist = list()
        by_course_enroll = statistics_query.by_course_enroll(date)
        for course_id, org, new_enroll_cnt, new_unenroll_cnt, all_enroll_cnt, all_unenroll_cnt, half_cnt, cert_cnt in by_course_enroll:
            row = tuple()
            # 0
            row += (get_value_from_dict(course_order, course_id, 99999),)
            row += (get_value_from_dict(dic_univ, org),)
            row += (get_value_from_dict(course_classfys, course_id, ''),)
            row += (get_value_from_dict(course_middle_classfys, course_id, ''),)
            row += (get_value_from_dict(course_names, course_id),)
            row += (get_value_from_dict(course_video, course_id),)
            row += (get_value_from_dict(course_effort, course_id),)
            row += (get_value_from_dict(course_week, course_id),)

            effort = get_value_from_dict(course_effort, course_id, None)
            week = get_value_from_dict(course_week, course_id, None)

            if effort and week:
                if effort.find(':') > 0:
                    hh = effort.split(':')[0]
                    mm = effort.split(':')[1]
                    row += (str((int(hh) * int(week)) + (int(mm) * int(week)) / 60) + ':' + ("%02d" % ((int(mm) * int(week)) % 60)),)
                else:
                    row += ('-',)
            else:
                row += ('-',)
            row += (org,)

            # 10
            row += (course_id.split('+')[1],)
            row += (course_id.split('+')[2],)
            row += (get_value_from_dict(course_state, course_id),)
            row += (get_value_from_dict(course_creates, course_id),)
            row += (get_value_from_dict(course_enroll_starts, course_id),)
            row += (get_value_from_dict(course_enroll_ends, course_id),)
            row += (get_value_from_dict(course_starts, course_id),)
            row += (get_value_from_dict(course_ends, course_id),)
            row += (get_value_from_dict(course_cert_date, course_id, ''),)
            row += (new_enroll_cnt,)
            row += (new_unenroll_cnt,)

            # 20
            row += (all_enroll_cnt,)
            row += (all_unenroll_cnt,)
            row += (all_enroll_cnt - all_unenroll_cnt,)

            # over 50% cert target
            row += (half_cnt if course_id in course_cert_date else '',)
            # certed target
            row += (cert_cnt if course_id in course_cert_date else '',)
            # course update date
            row += (get_value_from_dict(course_edited, course_id),)

            sortlist.append(row)

            print row

        # print 'course_order:', course_order
        sortlist.sort(key=itemgetter(0, 16, 4))
        # sortlist.sort(key=lambda order, cert, created: )

        start_row = 4
        for course_info in sortlist:

            # print 'course_info --- s'
            # print course_info
            # print 'course_info --- e'

            # order 값 제거
            course_info = course_info[1:]

            start_char = 65
            for idx in range(0, len(course_info)):
                ws3[chr(start_char) + str(start_row)] = course_info[idx]
                style_base(ws3[chr(start_char) + str(start_row)])
                start_char += 1
            start_row += 1

        # :SHEET 4
        # ------------------------> by_course_demographic

        sortlist = list()
        by_course_demographic = statistics_query.by_course_demographics(date)
        for course_id, org, \
            male_1, female_1, etc_1, no_gender1, age1_1, age2_1, age3_1, age4_1, age5_1, age6_1, edu1_1, edu2_1, edu3_1, edu4_1, edu5_1, edu6_1, edu7_1, edu8_1, edu9_1, allcnt_1, \
            male_2, female_2, etc_2, no_gender2, age1_2, age2_2, age3_2, age4_2, age5_2, age6_2, edu1_2, edu2_2, edu3_2, edu4_2, edu5_2, edu6_2, edu7_2, edu8_2, edu9_2, allcnt_2, \
            male_3, female_3, etc_3, no_gender3, age1_3, age2_3, age3_3, age4_3, age5_3, age6_3, edu1_3, edu2_3, edu3_3, edu4_3, edu5_3, edu6_3, edu7_3, edu8_3, edu9_3, allcnt_3 \
                in by_course_demographic:
            row = tuple()

            # 정렬을 위한 변수. 정렬 후 해당 내용은 사용하지 않음. s
            row += (get_value_from_dict(course_order, course_id, 99999),)
            row += (get_value_from_dict(course_starts, course_id, ''),)
            row += (get_value_from_dict(course_names, course_id, ''),)
            # e

            row += (get_value_from_dict(dic_univ, org),)
            row += (get_value_from_dict(course_classfys, course_id, ''),)
            row += (get_value_from_dict(course_middle_classfys, course_id, ''),)
            row += (get_value_from_dict(course_names, course_id),)
            row += (get_value_from_dict(course_video, course_id),)
            row += (get_value_from_dict(course_effort, course_id),)
            row += (get_value_from_dict(course_week, course_id),)

            effort = get_value_from_dict(course_effort, course_id, None)
            week = get_value_from_dict(course_week, course_id, None)

            if effort and week:
                if effort.find(':') > 0:
                    hh = effort.split(':')[0]
                    mm = effort.split(':')[1]
                    row += (str((int(hh) * int(week)) + (int(mm) * int(week)) / 60) + ':' + ("%02d" % ((int(mm) * int(week)) % 60)),)
                else:
                    row += ('-',)
            else:
                row += ('-',)

            row += (org,)
            row += (course_id.split('+')[1],)
            row += (course_id.split('+')[2],)

            row += (male_1,)
            row += (female_1,)
            row += (etc_1,)
            row += (no_gender1,)
            row += (age1_1,)
            row += (age2_1,)
            row += (age3_1,)
            row += (age4_1,)
            row += (age5_1,)
            row += (age6_1,)
            row += (edu1_1,)
            row += (edu2_1,)
            row += (edu3_1,)
            row += (edu4_1,)
            row += (edu5_1,)
            row += (edu6_1,)
            row += (edu7_1,)
            row += (edu8_1,)
            row += (edu9_1,)
            row += (allcnt_1,)

            row += (male_2 if course_id in course_cert_date else '-',)
            row += (female_2 if course_id in course_cert_date else '-',)
            row += (etc_2 if course_id in course_cert_date else '-',)
            row += (no_gender2 if course_id in course_cert_date else '-',)
            row += (age1_2 if course_id in course_cert_date else '-',)
            row += (age2_2 if course_id in course_cert_date else '-',)
            row += (age3_2 if course_id in course_cert_date else '-',)
            row += (age4_2 if course_id in course_cert_date else '-',)
            row += (age5_2 if course_id in course_cert_date else '-',)
            row += (age6_2 if course_id in course_cert_date else '-',)
            row += (edu1_2 if course_id in course_cert_date else '-',)
            row += (edu2_2 if course_id in course_cert_date else '-',)
            row += (edu3_2 if course_id in course_cert_date else '-',)
            row += (edu4_2 if course_id in course_cert_date else '-',)
            row += (edu5_2 if course_id in course_cert_date else '-',)
            row += (edu6_2 if course_id in course_cert_date else '-',)
            row += (edu7_2 if course_id in course_cert_date else '-',)
            row += (edu8_2 if course_id in course_cert_date else '-',)
            row += (edu9_2 if course_id in course_cert_date else '-',)
            row += (allcnt_2 if course_id in course_cert_date else '-',)

            row += (male_3 if course_id in course_cert_date else '-',)
            row += (female_3 if course_id in course_cert_date else '-',)
            row += (etc_3 if course_id in course_cert_date else '-',)
            row += (no_gender3 if course_id in course_cert_date else '-',)
            row += (age1_3 if course_id in course_cert_date else '-',)
            row += (age2_3 if course_id in course_cert_date else '-',)
            row += (age3_3 if course_id in course_cert_date else '-',)
            row += (age4_3 if course_id in course_cert_date else '-',)
            row += (age5_3 if course_id in course_cert_date else '-',)
            row += (age6_3 if course_id in course_cert_date else '-',)
            row += (edu1_3 if course_id in course_cert_date else '-',)
            row += (edu2_3 if course_id in course_cert_date else '-',)
            row += (edu3_3 if course_id in course_cert_date else '-',)
            row += (edu4_3 if course_id in course_cert_date else '-',)
            row += (edu5_3 if course_id in course_cert_date else '-',)
            row += (edu6_3 if course_id in course_cert_date else '-',)
            row += (edu7_3 if course_id in course_cert_date else '-',)
            row += (edu8_3 if course_id in course_cert_date else '-',)
            row += (edu9_3 if course_id in course_cert_date else '-',)
            row += (allcnt_3 if course_id in course_cert_date else '-',)

            sortlist.append(row)

        sortlist.sort(key=itemgetter(0, 1, 2))

        start_row = 4
        for course_info in sortlist:
            course_info = course_info[3:]

            start_char = 65
            for idx in range(0, len(course_info)):
                if start_char > 116:
                    ws4['B' + chr(start_char - 52) + str(start_row)] = course_info[idx]
                    style_base(ws4['B' + chr(start_char - 52) + str(start_row)])

                    start_char += 1
                elif start_char > 90:
                    ws4['A' + chr(start_char - 26) + str(start_row)] = course_info[idx]
                    style_base(ws4['A' + chr(start_char - 26) + str(start_row)])

                    start_char += 1
                else:
                    ws4[chr(start_char) + str(start_row)] = course_info[idx]
                    style_base(ws4[chr(start_char) + str(start_row)])

                    start_char += 1

            start_row += 1

        wb.save(save_path)
    return HttpResponse('/manage/home/static/excel/' + save_name, content_type='application/vnd.ms-excel')


def get_value_from_dict(dict, key, default=None):
    return dict[key] if key in dict else default
